"""
Retail order data cleaning pipeline - Part 1 Capstone.
Reads raw_orders.xlsx, produces cleaned_orders.xlsx, quality report, pivot summary, and screenshots.
"""

import re
from datetime import datetime
from pathlib import Path

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
import pandas as pd
from dateutil import parser
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

BASE = Path(__file__).resolve().parent
DATA_DIR = BASE / "data"
OUTPUT_DIR = BASE / "outputs"
SCREENSHOT_DIR = BASE / "screenshots"

RAW_PATH = DATA_DIR / "raw_orders.xlsx"
CLEANED_PATH = DATA_DIR / "cleaned_orders.xlsx"
QUALITY_PATH = OUTPUT_DIR / "data_quality_report.xlsx"
PIVOT_PATH = OUTPUT_DIR / "pivot_summary.xlsx"

TEXT_COLS = [
    "customer_name", "segment", "region", "state", "city",
    "category", "sub_category", "ship_mode", "payment_status", "order_status",
]

CATEGORY_MAP = {
    "office supplies": "Office Supplies",
    "furniture": "Furniture",
    "technology": "Technology",
}

SEGMENT_MAP = {
    "consumer": "Consumer",
    "small business": "Small Business",
    "corporate": "Corporate",
    "home office": "Home Office",
}

REGION_MAP = {
    "north": "North",
    "south": "South",
    "east": "East",
    "west": "West",
}

SHIP_MODE_MAP = {
    "standard class": "Standard Class",
    "second class": "Second Class",
    "first class": "First Class",
    "same day": "Same Day",
}

PAYMENT_MAP = {
    "paid": "Paid",
    "pending": "Pending",
    "refunded": "Refunded",
    "failed": "Failed",
}

ORDER_STATUS_MAP = {
    "completed": "Completed",
    "cancelled": "Cancelled",
    "returned": "Returned",
}

MAX_DISCOUNT = 1.0  # 100% - discounts above this are invalid


def clean_text(value):
    """TRIM + remove extra internal spaces + strip special chars."""
    if pd.isna(value):
        return value
    s = str(value)
    s = re.sub(r"[^\w\s\-&]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s if s else np.nan


def title_case_words(s):
    if pd.isna(s):
        return s
    return " ".join(w.capitalize() for w in str(s).split())


def standardize_categorical(value, mapping):
    if pd.isna(value):
        return value
    key = re.sub(r"\s+", " ", str(value).strip().lower())
    return mapping.get(key, title_case_words(value))


def parse_date(value):
    if pd.isna(value):
        return pd.NaT
    s = str(value).strip()
    if s.lower() in {"", "nan", "none", "n/a", "tbd", "invalid", "na"}:
        return pd.NaT
    if not re.search(r"\d", s):
        return pd.NaT
    try:
        return parser.parse(s, dayfirst=False)
    except (ValueError, TypeError, OverflowError):
        try:
            return parser.parse(s, dayfirst=True)
        except (ValueError, TypeError, OverflowError):
            return pd.NaT


def parse_discount(value, qty, unit_price, sales):
    """Return (cleaned_discount, issue_flag)."""
    if pd.isna(value):
        if pd.notna(qty) and pd.notna(unit_price) and pd.notna(sales):
            return 0.0, "missing_filled_zero"
        return np.nan, "missing_invalid"
    s = str(value).strip()
    if s.lower() in {"", "nan", "none", "n/a"}:
        if pd.notna(qty) and pd.notna(unit_price) and pd.notna(sales):
            return 0.0, "missing_filled_zero"
        return np.nan, "missing_invalid"
    try:
        if "%" in s:
            num = float(s.replace("%", "").strip()) / 100.0
        else:
            num = float(s)
    except ValueError:
        return np.nan, "invalid_format"
    if num < 0:
        return num, "negative"
    if num > MAX_DISCOUNT:
        return num, "above_range"
    return num, "ok"


def style_sheet(ws, header_fill="366092"):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)


def save_table_screenshot(df, title, path, nrows=12, ncols=10):
    """Render a dataframe preview as PNG for assignment screenshots."""
    preview = df.iloc[:nrows, :ncols].copy()
    fig, ax = plt.subplots(figsize=(16, 6))
    ax.axis("off")
    ax.set_title(title, fontsize=14, fontweight="bold", pad=12)

    table_data = []
    for _, row in preview.iterrows():
        table_data.append([str(v)[:28] if pd.notna(v) else "" for v in row])

    tbl = ax.table(
        cellText=table_data,
        colLabels=[str(c)[:20] for c in preview.columns],
        loc="center",
        cellLoc="left",
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(7)
    tbl.scale(1, 1.4)
    for (r, c), cell in tbl.get_celld().items():
        if r == 0:
            cell.set_facecolor("#366092")
            cell.set_text_props(color="white", fontweight="bold")
        elif r % 2 == 0:
            cell.set_facecolor("#f2f2f2")
    plt.tight_layout()
    fig.savefig(path, dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def save_pivot_screenshot(df, title, path, chart_type="bar"):
    fig, ax = plt.subplots(figsize=(12, 6))
    if chart_type == "bar" and len(df.columns) >= 2:
        x = df.iloc[:, 0].astype(str)
        y1 = pd.to_numeric(df.iloc[:, 1], errors="coerce")
        ax.bar(x, y1, color="#4472C4", label=str(df.columns[1]))
        if len(df.columns) >= 3:
            y2 = pd.to_numeric(df.iloc[:, 2], errors="coerce")
            ax.bar(x, y2, bottom=y1, color="#ED7D31", label=str(df.columns[2]))
            ax.legend()
        ax.set_xticklabels(x, rotation=45, ha="right")
    elif chart_type == "line":
        x = df.iloc[:, 0].astype(str)
        y = pd.to_numeric(df.iloc[:, 1], errors="coerce")
        ax.plot(x, y, marker="o", color="#4472C4", linewidth=2)
        ax.set_xticklabels(x, rotation=45, ha="right")
    ax.set_title(title, fontsize=13, fontweight="bold")
    ax.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    fig.savefig(path, dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def main():
    DATA_DIR.mkdir(exist_ok=True)
    OUTPUT_DIR.mkdir(exist_ok=True)
    SCREENSHOT_DIR.mkdir(exist_ok=True)

    raw = pd.read_excel(RAW_PATH)
    df = raw.copy()
    quality = {}

    # --- Task 2: Clean text fields ---
    for col in TEXT_COLS:
        df[col] = df[col].apply(clean_text)

    df["segment"] = df["segment"].apply(lambda x: standardize_categorical(x, SEGMENT_MAP))
    df["region"] = df["region"].apply(lambda x: standardize_categorical(x, REGION_MAP))
    df["category"] = df["category"].apply(lambda x: standardize_categorical(x, CATEGORY_MAP))
    df["ship_mode"] = df["ship_mode"].apply(lambda x: standardize_categorical(x, SHIP_MODE_MAP))
    df["payment_status"] = df["payment_status"].apply(lambda x: standardize_categorical(x, PAYMENT_MAP))
    df["order_status"] = df["order_status"].apply(lambda x: standardize_categorical(x, ORDER_STATUS_MAP))
    df["customer_name"] = df["customer_name"].apply(title_case_words)
    df["state"] = df["state"].apply(title_case_words)
    df["city"] = df["city"].apply(title_case_words)
    df["sub_category"] = df["sub_category"].apply(title_case_words)

    # --- Task 3: Dates ---
    df["order_date"] = df["order_date"].apply(parse_date)
    df["ship_date"] = df["ship_date"].apply(parse_date)
    df["shipping_delay_days"] = (df["ship_date"] - df["order_date"]).dt.days

    date_issues = []
    for idx, row in df.iterrows():
        flags = []
        if pd.isna(row["order_date"]):
            flags.append("missing_order_date")
        if pd.isna(row["ship_date"]):
            flags.append("missing_ship_date")
        if pd.notna(row["order_date"]) and pd.notna(row["ship_date"]) and row["ship_date"] < row["order_date"]:
            flags.append("ship_before_order")
        if flags:
            date_issues.append({"order_id": row["order_id"], "issues": "; ".join(flags)})

    quality["date_issues"] = pd.DataFrame(date_issues) if date_issues else pd.DataFrame(
        columns=["order_id", "issues"]
    )

    # --- Task 4: Duplicates ---
    exact_dup_count = int(df.duplicated().sum())
    exact_dup_removed = exact_dup_count  # we remove all but first

    dup_id_mask = df["order_id"].duplicated(keep=False)
    dup_id_count = int(df.loc[dup_id_mask, "order_id"].nunique())

    conflicting_ids = []
    flagged_dup_rows = []
    for oid, grp in df.groupby("order_id"):
        if len(grp) <= 1:
            continue
        unique_rows = grp.drop_duplicates()
        if len(unique_rows) > 1:
            conflicting_ids.append(oid)
            for idx in grp.index:
                flagged_dup_rows.append(idx)

    # Remove exact duplicates (keep first occurrence)
    before_count = len(df)
    df = df.drop_duplicates(keep="first").reset_index(drop=True)
    records_removed = before_count - len(df)

    dup_summary = pd.DataFrame([
        {"Metric": "Exact duplicate rows found", "Count": exact_dup_count},
        {"Metric": "Duplicate order_id groups found", "Count": dup_id_count},
        {"Metric": "Conflicting duplicate order_id groups", "Count": len(conflicting_ids)},
        {"Metric": "Records removed (exact duplicates only)", "Count": records_removed},
        {"Metric": "Records flagged for review (conflicting order_id)", "Count": len(flagged_dup_rows)},
        {"Metric": "Logic", "Count": (
            "Exact duplicate rows were removed keeping the first occurrence. "
            "Conflicting order_id records were retained and flagged in data_quality_flag "
            "for manual review rather than silently deleted."
        )},
    ])

    # --- Task 5 & 6: Business rules + calculated columns ---
    discount_results = df.apply(
        lambda r: parse_discount(r["discount"], r["quantity"], r["unit_price"], r["sales"]),
        axis=1,
    )
    df["cleaned_discount"] = [d[0] for d in discount_results]
    df["discount_issue"] = [d[1] for d in discount_results]

    # Missing region / ship_mode
    missing_region = df["region"].isna()
    missing_ship = df["ship_mode"].isna()
    df.loc[missing_region, "region"] = "Unknown"
    df.loc[missing_ship, "ship_mode"] = "Unknown"

    df["calculated_sales"] = df["quantity"] * df["unit_price"] * (1 - df["cleaned_discount"].fillna(0))
    df["calculated_profit"] = df["calculated_sales"] - df["cost"]
    df["profit_margin"] = np.where(
        df["calculated_sales"] != 0,
        df["calculated_profit"] / df["calculated_sales"],
        np.nan,
    )
    df["order_month"] = df["order_date"].dt.month
    df["order_year"] = df["order_date"].dt.year

    sales_mismatch = abs(df["calculated_sales"] - df["sales"]) > 1.0

    def assign_quality_flag(row):
        flags = []
        if row["order_id"] in conflicting_ids:
            flags.append("duplicate_order_id_conflict")
        if row["discount_issue"] in ("negative", "above_range", "invalid_format", "missing_invalid"):
            flags.append(f"discount_{row['discount_issue']}")
        elif row["discount_issue"] == "missing_filled_zero":
            flags.append("discount_missing_filled_zero")
        if pd.isna(row["order_date"]):
            flags.append("missing_order_date")
        if pd.isna(row["ship_date"]):
            flags.append("missing_ship_date")
        if pd.notna(row["order_date"]) and pd.notna(row["ship_date"]) and row["ship_date"] < row["order_date"]:
            flags.append("ship_before_order")
        if pd.isna(row["region"]) or row["region"] == "Unknown":
            if "region" in str(row.get("_was_missing_region", "")):
                pass
        if missing_region.reindex(df.index, fill_value=False).get(row.name, False) if hasattr(row, "name") else False:
            flags.append("region_was_missing")
        if sales_mismatch.loc[row.name] if row.name in sales_mismatch.index else False:
            flags.append("sales_calc_mismatch")

        invalid_status = (
            row["order_status"] in ("Cancelled",)
            and row["payment_status"] == "Failed"
        )
        if invalid_status:
            flags.append("cancelled_and_failed")

        if any(f.startswith("discount_negative") or f.startswith("discount_above") or
               f.startswith("discount_invalid") or f == "missing_order_date" or
               f == "missing_ship_date" for f in flags):
            return "invalid", "; ".join(flags) if flags else "clean"
        if flags:
            return "warning", "; ".join(flags)
        return "clean", "none"

    # Recompute flags with index-aware helpers
    region_was_missing = missing_region.reindex(df.index, fill_value=False)
    ship_was_missing = missing_ship.reindex(df.index, fill_value=False)
    conflict_set = set(conflicting_ids)

    flag_results = []
    for idx, row in df.iterrows():
        flags = []
        if row["order_id"] in conflict_set:
            flags.append("duplicate_order_id_conflict")
        if row["discount_issue"] in ("negative", "above_range", "invalid_format", "missing_invalid"):
            flags.append(f"discount_{row['discount_issue']}")
        elif row["discount_issue"] == "missing_filled_zero":
            flags.append("discount_missing_filled_zero")
        if pd.isna(row["order_date"]):
            flags.append("missing_order_date")
        if pd.isna(row["ship_date"]):
            flags.append("missing_ship_date")
        if pd.notna(row["order_date"]) and pd.notna(row["ship_date"]) and row["ship_date"] < row["order_date"]:
            flags.append("ship_before_order")
        if region_was_missing.get(idx, False):
            flags.append("region_was_missing")
        if ship_was_missing.get(idx, False):
            flags.append("ship_mode_was_missing")
        if sales_mismatch.loc[idx]:
            flags.append("sales_calc_mismatch")

        severe = any(
            f.startswith("discount_negative") or f.startswith("discount_above") or
            f.startswith("discount_invalid") or f in ("missing_order_date", "missing_ship_date")
            for f in flags
        )
        if severe:
            flag_results.append(("invalid", "; ".join(flags)))
        elif flags:
            flag_results.append(("warning", "; ".join(flags)))
        else:
            flag_results.append(("clean", "none"))

    df["data_quality_flag"] = [f[0] for f in flag_results]
    df["quality_flag_detail"] = [f[1] for f in flag_results]

    # Format dates for Excel output
    export_df = df.copy()
    export_df["order_date"] = export_df["order_date"].dt.strftime("%Y-%m-%d")
    export_df["ship_date"] = export_df["ship_date"].dt.strftime("%Y-%m-%d")
    export_df["order_date"] = export_df["order_date"].replace("NaT", "")
    export_df["ship_date"] = export_df["ship_date"].replace("NaT", "")

    export_df.to_excel(CLEANED_PATH, index=False, sheet_name="Cleaned Orders")

    # --- Task 7: Data Quality Report ---
    missing_summary = pd.DataFrame({
        "Column": TEXT_COLS + ["order_date", "ship_date", "discount", "quantity", "unit_price"],
        "Missing Count (Raw)": [int(raw[c].isna().sum()) if c in raw.columns else 0
                                for c in TEXT_COLS + ["order_date", "ship_date", "discount", "quantity", "unit_price"]],
        "Missing Count (Cleaned)": [
            int((export_df[c].isna() | (export_df[c].astype(str).str.strip() == "")).sum())
            if c in export_df.columns else 0
            for c in TEXT_COLS + ["order_date", "ship_date", "discount", "quantity", "unit_price"]
        ],
        "Action Taken": [
            "TRIM + standardize case/spacing" if c in TEXT_COLS else
            "Parsed to YYYY-MM-DD; invalid flagged" if c in ("order_date", "ship_date") else
            "Filled as 0 when sales fields valid; else flagged" if c == "discount" else
            "No action needed"
            for c in TEXT_COLS + ["order_date", "ship_date", "discount", "quantity", "unit_price"]
        ],
    })

    invalid_discount = df[df["discount_issue"].isin(["negative", "above_range", "invalid_format", "missing_invalid"])][
        ["order_id", "discount", "cleaned_discount", "discount_issue"]
    ].copy()

    date_summary = pd.DataFrame([
        {"Issue": "Missing order_date", "Count": int(df["order_date"].isna().sum())},
        {"Issue": "Missing ship_date", "Count": int(df["ship_date"].isna().sum())},
        {"Issue": "Ship date before order date", "Count": int((df["shipping_delay_days"] < 0).sum())},
        {"Issue": "Discount filled as 0 (was missing)", "Count": int((df["discount_issue"] == "missing_filled_zero").sum())},
    ])

    order_status_summary = pd.DataFrame([
        {"order_status": k, "payment_status": v, "count": int(c)}
        for (k, v), c in df.groupby(["order_status", "payment_status"]).size().items()
    ]).sort_values("count", ascending=False)

    calc_mismatch = df[sales_mismatch][
        ["order_id", "quantity", "unit_price", "cleaned_discount", "sales", "calculated_sales", "profit", "calculated_profit"]
    ].copy()
    calc_mismatch["sales_diff"] = calc_mismatch["calculated_sales"] - calc_mismatch["sales"]
    calc_mismatch["profit_diff"] = calc_mismatch["calculated_profit"] - calc_mismatch["profit"]

    final_counts = pd.DataFrame([
        {"Flag": "clean", "Count": int((df["data_quality_flag"] == "clean").sum())},
        {"Flag": "warning", "Count": int((df["data_quality_flag"] == "warning").sum())},
        {"Flag": "invalid", "Count": int((df["data_quality_flag"] == "invalid").sum())},
        {"Flag": "Total records after dedup", "Count": len(df)},
    ])

    with pd.ExcelWriter(QUALITY_PATH, engine="openpyxl") as writer:
        missing_summary.to_excel(writer, sheet_name="Missing Values", index=False)
        dup_summary.to_excel(writer, sheet_name="Duplicate Summary", index=False)
        invalid_discount.to_excel(writer, sheet_name="Invalid Discounts", index=False)
        date_summary.to_excel(writer, sheet_name="Date Issues", index=False)
        order_status_summary.to_excel(writer, sheet_name="Order Status", index=False)
        calc_mismatch.to_excel(writer, sheet_name="Calc Mismatches", index=False)
        final_counts.to_excel(writer, sheet_name="Final Counts", index=False)

    wb = load_workbook(QUALITY_PATH)
    for ws in wb.worksheets:
        style_sheet(ws)
    wb.save(QUALITY_PATH)

    # --- Task 8: Pivot Summary ---
    completed_sales = df[
        (df["order_status"] == "Completed") &
        (df["payment_status"] == "Paid")
    ].copy()

    pivot_region = (
        completed_sales.groupby("region", dropna=False)
        .agg(Sales=("calculated_sales", "sum"), Profit=("calculated_profit", "sum"))
        .reset_index()
        .sort_values("Sales", ascending=False)
    )

    pivot_category = (
        completed_sales.groupby(["category", "sub_category"], dropna=False)
        .agg(Sales=("calculated_sales", "sum"), Profit=("calculated_profit", "sum"))
        .reset_index()
        .sort_values("Profit", ascending=False)
    )

    pivot_ship = (
        df.groupby("ship_mode", dropna=False)
        .agg(Order_Count=("order_id", "count"))
        .reset_index()
        .sort_values("Order_Count", ascending=False)
    )

    pivot_margin = (
        completed_sales.groupby("segment", dropna=False)
        .agg(
            Avg_Profit_Margin=("profit_margin", "mean"),
            Total_Sales=("calculated_sales", "sum"),
        )
        .reset_index()
        .sort_values("Avg_Profit_Margin", ascending=False)
    )

    problem_orders = df[
        (df["order_status"].isin(["Cancelled", "Returned"])) |
        (df["payment_status"].isin(["Refunded", "Failed"]))
    ]
    pivot_problems = (
        problem_orders.groupby(["region", "order_status", "payment_status"], dropna=False)
        .agg(Order_Count=("order_id", "count"))
        .reset_index()
        .sort_values("Order_Count", ascending=False)
    )

    pivot_monthly = (
        completed_sales.groupby(["order_year", "order_month"], dropna=False)
        .agg(Monthly_Sales=("calculated_sales", "sum"))
        .reset_index()
        .sort_values(["order_year", "order_month"])
    )
    pivot_monthly["Year_Month"] = (
        pivot_monthly["order_year"].astype("Int64").astype(str) + "-" +
        pivot_monthly["order_month"].astype("Int64").astype(str).str.zfill(2)
    )

    refunded_summary = df[df["payment_status"] == "Refunded"].groupby("region").agg(
        Refunded_Orders=("order_id", "count"),
        Refunded_Sales=("calculated_sales", "sum"),
    ).reset_index().sort_values("Refunded_Sales", ascending=False)

    with pd.ExcelWriter(PIVOT_PATH, engine="openpyxl") as writer:
        pivot_region.to_excel(writer, sheet_name="Sales by Region", index=False)
        pivot_category.to_excel(writer, sheet_name="Sales by Category", index=False)
        pivot_ship.to_excel(writer, sheet_name="Orders by Ship Mode", index=False)
        pivot_margin.to_excel(writer, sheet_name="Margin by Segment", index=False)
        pivot_problems.to_excel(writer, sheet_name="Problem Orders", index=False)
        pivot_monthly[["Year_Month", "Monthly_Sales"]].to_excel(writer, sheet_name="Monthly Trend", index=False)
        refunded_summary.to_excel(writer, sheet_name="Refunded by Region", index=False)

    wb2 = load_workbook(PIVOT_PATH)
    for ws in wb2.worksheets:
        style_sheet(ws)
        if ws.title == "Sales by Region":
            ws.auto_filter.ref = ws.dimensions
        if ws.title == "Sales by Category":
            ws.auto_filter.ref = ws.dimensions
    wb2.save(PIVOT_PATH)

    # --- Screenshots ---
    raw_preview = pd.read_excel(RAW_PATH)
    save_table_screenshot(raw_preview, "Raw Orders Dataset - Preview", SCREENSHOT_DIR / "raw_data_preview.png")

    cleaned_preview_cols = [
        "order_id", "order_date", "ship_date", "customer_name", "region", "category",
        "quantity", "unit_price", "cleaned_discount", "calculated_sales",
        "calculated_profit", "profit_margin", "shipping_delay_days", "data_quality_flag",
    ]
    save_table_screenshot(
        export_df[cleaned_preview_cols],
        "Cleaned Orders - With Calculated Columns",
        SCREENSHOT_DIR / "cleaned_data_preview.png",
        ncols=14,
    )

    save_pivot_screenshot(
        pivot_region.head(10),
        "Pivot: Sales & Profit by Region (Completed + Paid, sorted by Sales desc)",
        SCREENSHOT_DIR / "pivot_summary_1.png",
        chart_type="bar",
    )

    save_pivot_screenshot(
        pivot_monthly[["Year_Month", "Monthly_Sales"]].tail(18),
        "Pivot: Monthly Sales Trend (Completed + Paid orders)",
        SCREENSHOT_DIR / "pivot_summary_2.png",
        chart_type="line",
    )

    # Store stats for log generation
    stats = {
        "raw_rows": len(raw),
        "cleaned_rows": len(df),
        "exact_dup_count": exact_dup_count,
        "dup_id_count": dup_id_count,
        "conflicting_ids": conflicting_ids,
        "records_removed": records_removed,
        "missing_region": int(missing_region.sum()),
        "missing_ship": int(missing_ship.sum()),
        "invalid_discount_count": len(invalid_discount),
        "date_issue_count": len(date_issues),
        "ship_before_order": int((df["shipping_delay_days"] < 0).sum()),
        "sales_mismatch_count": int(sales_mismatch.sum()),
        "clean_count": int((df["data_quality_flag"] == "clean").sum()),
        "warning_count": int((df["data_quality_flag"] == "warning").sum()),
        "invalid_count": int((df["data_quality_flag"] == "invalid").sum()),
        "pivot_region": pivot_region,
        "pivot_category": pivot_category.head(5),
        "pivot_monthly": pivot_monthly,
        "completed_sales_total": completed_sales["calculated_sales"].sum(),
        "refunded_total": df[df["payment_status"] == "Refunded"]["calculated_sales"].sum(),
    }
    return stats


if __name__ == "__main__":
    stats = main()
    print("Cleaning complete.")
    for k, v in stats.items():
        if not isinstance(v, (pd.DataFrame, list)):
            print(f"  {k}: {v}")
